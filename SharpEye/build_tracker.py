"""
Build SharpEye Outreach Tracker workbook.
Generates SharpEye_Outreach_Tracker.xlsx with 6 tabs.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from copy import copy

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=11)
BODY_ALIGN = Alignment(vertical="top", wrap_text=False)

ALT_FILL_EVEN = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ALT_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Conditional formatting fills
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(color="9C5700")

# Section header style for Templates and Scripts tabs
SECTION_FONT = Font(name="Calibri", bold=True, size=13, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")


def style_header_row(ws, num_cols, row=1):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def apply_body_style(ws, start_row, end_row, num_cols):
    """Apply alternating row colors and body font to a data range."""
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL_EVEN if (r - start_row) % 2 == 0 else ALT_FILL_ODD
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.fill = fill
            cell.border = THIN_BORDER


def add_auto_filter_and_freeze(ws, num_cols, header_row=1):
    """Freeze header row and enable auto-filter."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"


def add_status_dropdown(ws, col_letter, start_row, end_row, options):
    """Add a data validation dropdown for a status column."""
    dv = DataValidation(
        type="list",
        formula1='"' + ','.join(options) + '"',
        allow_blank=True,
    )
    dv.error = "Please select a value from the dropdown."
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


# ---------------------------------------------------------------------------
# Workbook creation
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()

# ===== TAB 1: Lead List =====
ws1 = wb.active
ws1.title = "Lead List"

LEAD_COLS = [
    "Lead ID", "Company Name", "Contact First Name", "Contact Last Name",
    "Title", "Email", "Phone", "Company City", "Company State",
    "Industry", "Lead Score (1-5)", "Status", "Assigned Campaign", "Notes",
]
NUM_LEAD_COLS = len(LEAD_COLS)

for c, name in enumerate(LEAD_COLS, 1):
    ws1.cell(row=1, column=c, value=name)

# Pre-fill 30 sample rows so columns render nicely
SAMPLE_DATA = 30
for r in range(2, 2 + SAMPLE_DATA):
    ws1.cell(row=r, column=1, value=r - 1)  # Lead ID

style_header_row(ws1, NUM_LEAD_COLS)
apply_body_style(ws1, 2, 2 + SAMPLE_DATA, NUM_LEAD_COLS)
add_auto_filter_and_freeze(ws1, NUM_LEAD_COLS)

# Data validation: Lead Score 1-5
dv_score = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
dv_score.error = "Score must be 1–5"
ws1.add_data_validation(dv_score)
dv_score.add(f"K2:K{2 + SAMPLE_DATA}")

# Status dropdown
status_options = [
    "New", "Emailed", "Called", "Replied", "Meeting Booked",
    "Quoted", "Won", "Lost",
]
add_status_dropdown(ws1, "L", 2, 2 + SAMPLE_DATA, status_options)

# Conditional formatting: Status column
ws1.conditional_formatting.add(
    f"L2:L{2 + SAMPLE_DATA}",
    CellIsRule(operator="equal", formula=['"Lost"'], fill=RED_FILL, font=RED_FONT),
)
ws1.conditional_formatting.add(
    f"L2:L{2 + SAMPLE_DATA}",
    CellIsRule(operator="equal", formula=['"Won"'], fill=GREEN_FILL, font=GREEN_FONT),
)
ws1.conditional_formatting.add(
    f"L2:L{2 + SAMPLE_DATA}",
    CellIsRule(
        operator="equal", formula=['"Meeting Booked"'], fill=YELLOW_FILL, font=YELLOW_FONT
    ),
)

# Set column widths
_col_widths_1 = [10, 28, 18, 20, 24, 32, 16, 20, 14, 22, 14, 16, 22, 36]
for i, w in enumerate(_col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ===== TAB 2: Email Campaigns =====
ws2 = wb.create_sheet("Email Campaigns")

EMAIL_COLS = [
    "Lead ID", "Contact Name", "Company", "Campaign Name",
    "Email Sent Date", "Opened?", "Clicked?", "Replied?",
    "Reply Notes",
    "Follow-up #1 Date", "Follow-up #2 Date", "Follow-up #3 Date",
    "Status",
]
NUM_EMAIL_COLS = len(EMAIL_COLS)

for c, name in enumerate(EMAIL_COLS, 1):
    ws2.cell(row=1, column=c, value=name)

# Pre-fill 30 rows with formulas for follow-up dates
for r in range(2, 2 + SAMPLE_DATA):
    ws2.cell(row=r, column=1, value=r - 1)
    # Follow-up #1 Date = Sent Date + 3 business days
    ws2.cell(row=r, column=10).value = (
        f'=IF(E{r}="","",WORKDAY(E{r},3))'
    )
    # Follow-up #2 Date = Follow-up #1 + 5 business days
    ws2.cell(row=r, column=11).value = (
        f'=IF(J{r}="","",WORKDAY(J{r},5))'
    )
    # Follow-up #3 Date = Follow-up #2 + 7 business days
    ws2.cell(row=r, column=12).value = (
        f'=IF(K{r}="","",WORKDAY(K{r},7))'
    )

style_header_row(ws2, NUM_EMAIL_COLS)
apply_body_style(ws2, 2, 2 + SAMPLE_DATA, NUM_EMAIL_COLS)
add_auto_filter_and_freeze(ws2, NUM_EMAIL_COLS)

# Yes/No dropdowns for Opened, Clicked, Replied
dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws2.add_data_validation(dv_yn)
for col_l in ["F", "G", "H"]:
    dv_yn.add(f"{col_l}2:{col_l}{2 + SAMPLE_DATA}")

# Email Status dropdown
email_status = ["Awaiting Reply", "Replied", "Opted Out", "Bounced"]
add_status_dropdown(ws2, "M", 2, 2 + SAMPLE_DATA, email_status)

# Conditional formatting: Status
for col in ["M"]:
    ws2.conditional_formatting.add(
        f"{col}2:{col}{2 + SAMPLE_DATA}",
        CellIsRule(operator="equal", formula=['"Replied"'], fill=GREEN_FILL, font=GREEN_FONT),
    )
    ws2.conditional_formatting.add(
        f"{col}2:{col}{2 + SAMPLE_DATA}",
        CellIsRule(operator="equal", formula=['"Opted Out"'], fill=RED_FILL, font=RED_FONT),
    )
    ws2.conditional_formatting.add(
        f"{col}2:{col}{2 + SAMPLE_DATA}",
        CellIsRule(operator="equal", formula=['"Bounced"'], fill=RED_FILL, font=RED_FONT),
    )

_col_widths_2 = [10, 22, 28, 26, 16, 12, 12, 12, 36, 18, 18, 18, 18]
for i, w in enumerate(_col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ===== TAB 3: Cold Calls =====
ws3 = wb.create_sheet("Cold Calls")

CALL_COLS = [
    "Lead ID", "Contact Name", "Company", "Phone Number",
    "Call Date", "Call Time", "Connected?",
    "Call Outcome", "Notes from Call",
    "Follow-up Task", "Follow-up Date",
]
NUM_CALL_COLS = len(CALL_COLS)

for c, name in enumerate(CALL_COLS, 1):
    ws3.cell(row=1, column=c, value=name)

SAMPLE_CALLS = 30
for r in range(2, 2 + SAMPLE_CALLS):
    ws3.cell(row=r, column=1, value=r - 1)

style_header_row(ws3, NUM_CALL_COLS)
apply_body_style(ws3, 2, 2 + SAMPLE_CALLS, NUM_CALL_COLS)
add_auto_filter_and_freeze(ws3, NUM_CALL_COLS)

# Connected dropdown
dv_conn = DataValidation(type="list", formula1='"Yes,No,VM Left"', allow_blank=True)
ws3.add_data_validation(dv_conn)
dv_conn.add(f"G2:G{2 + SAMPLE_CALLS}")

# Call Outcome dropdown
call_outcomes = [
    "No Answer", "Wrong Number", "Not Interested",
    "Call Back Later", "Interested", "Meeting Booked",
]
add_status_dropdown(ws3, "H", 2, 2 + SAMPLE_CALLS, call_outcomes)

# Conditional formatting: Call Outcome
ws3.conditional_formatting.add(
    f"H2:H{2 + SAMPLE_CALLS}",
    CellIsRule(
        operator="equal", formula=['"Meeting Booked"'], fill=YELLOW_FILL, font=YELLOW_FONT
    ),
)
ws3.conditional_formatting.add(
    f"H2:H{2 + SAMPLE_CALLS}",
    CellIsRule(
        operator="equal", formula=['"Not Interested"'], fill=RED_FILL, font=RED_FONT
    ),
)
ws3.conditional_formatting.add(
    f"H2:H{2 + SAMPLE_CALLS}",
    CellIsRule(
        operator="equal", formula=['"Interested"'], fill=GREEN_FILL, font=GREEN_FONT
    ),
)

_col_widths_3 = [10, 22, 28, 16, 14, 12, 14, 18, 40, 36, 16]
for i, w in enumerate(_col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ===== TAB 4: Dashboard =====
ws4 = wb.create_sheet("Dashboard")

# -- Section A: Summary KPIs --
SECTION_A_START = 1
dash_headers = ["Metric", "Value"]
for c, h in enumerate(dash_headers, 1):
    ws4.cell(row=SECTION_A_START, column=c, value=h)

metrics = [
    ("Total Leads", "=COUNTA('Lead List'!A2:A1001)"),
    ("Leads - New", "=COUNTIF('Lead List'!L2:L1001,\"New\")"),
    ("Leads - Emailed", "=COUNTIF('Lead List'!L2:L1001,\"Emailed\")"),
    ("Leads - Called", "=COUNTIF('Lead List'!L2:L1001,\"Called\")"),
    ("Leads - Replied", "=COUNTIF('Lead List'!L2:L1001,\"Replied\")"),
    ("Leads - Meeting Booked", "=COUNTIF('Lead List'!L2:L1001,\"Meeting Booked\")"),
    ("Leads - Quoted", "=COUNTIF('Lead List'!L2:L1001,\"Quoted\")"),
    ("Leads - Won", "=COUNTIF('Lead List'!L2:L1001,\"Won\")"),
    ("Leads - Lost", "=COUNTIF('Lead List'!L2:L1001,\"Lost\")"),
    ("", ""),
    ("--- Campaign Metrics ---", ""),
    ("Emails Sent This Month", "=SUMPRODUCT((MONTH('Email Campaigns'!E2:E1001)=MONTH(TODAY()))*(YEAR('Email Campaigns'!E2:E1001)=YEAR(TODAY()))*(LEN('Email Campaigns'!E2:E1001)>0))"),
    ("Open Rate %", "=IFERROR(SUMPRODUCT((MONTH('Email Campaigns'!E2:E1001)=MONTH(TODAY()))*('Email Campaigns'!F2:F1001=\"Yes\"))/MAX(1,SUMPRODUCT((MONTH('Email Campaigns'!E2:E1001)=MONTH(TODAY()))*(LEN('Email Campaigns'!E2:E1001)>0))),0)"),
    ("Reply Rate %", "=IFERROR(SUMPRODUCT((MONTH('Email Campaigns'!E2:E1001)=MONTH(TODAY()))*('Email Campaigns'!H2:H1001=\"Yes\"))/MAX(1,SUMPRODUCT((MONTH('Email Campaigns'!E2:E1001)=MONTH(TODAY()))*(LEN('Email Campaigns'!E2:E1001)>0))),0)"),
    ("Calls Made This Month", "=SUMPRODUCT((MONTH('Cold Calls'!E2:E1001)=MONTH(TODAY()))*(YEAR('Cold Calls'!E2:E1001)=YEAR(TODAY()))*(LEN('Cold Calls'!E2:E1001)>0))"),
    ("Meetings Booked (Calls)", "=COUNTIF('Cold Calls'!H2:H1001,\"Meeting Booked\")"),
    ("Meetings Booked (All)", "=COUNTIF('Lead List'!L2:L1001,\"Meeting Booked\")"),
]

for i, (label, formula) in enumerate(metrics, 2):
    ws4.cell(row=i, column=1, value=label)
    ws4.cell(row=i, column=2, value=formula)

style_header_row(ws4, 2, SECTION_A_START)
last_data_row = 2 + len(metrics) - 1
apply_body_style(ws4, 2, last_data_row, 2)

# -- Section B: Conversion Funnel --
FUNNEL_START = last_data_row + 3
ws4.cell(row=FUNNEL_START, column=1, value="--- Conversion Funnel ---")
ws4.cell(row=FUNNEL_START, column=1).font = Font(bold=True, size=12, color="1F4E79")

funnel_headers = ["Stage", "Count"]
for c, h in enumerate(funnel_headers, 1):
    ws4.cell(row=FUNNEL_START + 1, column=c, value=h)
style_header_row(ws4, 2, FUNNEL_START + 1)

funnel_stages = [
    ("Total Leads", "=COUNTA('Lead List'!A2:A1001)"),
    ("Emailed", "=COUNTIF('Lead List'!L2:L1001,\"Emailed\")"),
    ("Replied", "=COUNTIF('Lead List'!L2:L1001,\"Replied\")"),
    ("Meeting Booked", "=COUNTIF('Lead List'!L2:L1001,\"Meeting Booked\")"),
    ("Quoted", "=COUNTIF('Lead List'!L2:L1001,\"Quoted\")"),
    ("Won", "=COUNTIF('Lead List'!L2:L1001,\"Won\")"),
]

for i, (stage, formula) in enumerate(funnel_stages):
    r = FUNNEL_START + 2 + i
    ws4.cell(row=r, column=1, value=stage)
    ws4.cell(row=r, column=2, value=formula)
apply_body_style(ws4, FUNNEL_START + 2, FUNNEL_START + 2 + len(funnel_stages) - 1, 2)

# -- Section C: Hottest Leads (top 10 by stage/recency) --
HOT_START = FUNNEL_START + 2 + len(funnel_stages) + 2
ws4.cell(row=HOT_START, column=1, value="--- Top 10 Hottest Leads (by Stage) ---")
ws4.cell(row=HOT_START, column=1).font = Font(bold=True, size=12, color="1F4E79")

hot_headers = ["Lead ID", "Company", "Contact", "Email", "Status", "Notes"]
for c, h in enumerate(hot_headers, 1):
    ws4.cell(row=HOT_START + 1, column=c, value=h)
style_header_row(ws4, len(hot_headers), HOT_START + 1)

# Reference formulas to pull top leads (manual sort hint)
for i in range(1, 11):
    r = HOT_START + 1 + i
    ws4.cell(row=r, column=1, value=f"(Sort Lead List by Status/Recency)")
    ws4.cell(row=r, column=1).font = Font(italic=True, color="808080")

apply_body_style(ws4, HOT_START + 2, HOT_START + 11, len(hot_headers))

# -- Percentage formatting --
for r in range(2, last_data_row + 1):
    val = ws4.cell(row=r, column=2).value
    if val and "Rate" in str(val):
        ws4.cell(row=r, column=2).number_format = '0.0%'

ws4.column_dimensions["A"].width = 40
ws4.column_dimensions["B"].width = 18
ws4.freeze_panes = "A2"

# ===== TAB 5: Email Templates =====
ws5 = wb.create_sheet("Email Templates")

TEMPLATE_COLS = ["Template #", "Template Name", "Subject Line", "Body"]
for c, name in enumerate(TEMPLATE_COLS, 1):
    ws5.cell(row=1, column=c, value=name)
style_header_row(ws5, len(TEMPLATE_COLS))

templates = [
    (
        1,
        "Initial Outreach",
        "Sharp-Eye Precision Parts — supporting {{Company}}'s machining needs",
        (
            "Hi {{FirstName}},\n\n"
            "I'm August Cox, country manager at Sharp-Eye Precision Parts Co., Ltd. "
            "We specialize in CNC machining turned parts — swiss-type automatic lathe work "
            "is what we do best.\n\n"
            "I've been looking at what {{Company}} does in the {{Industry}} space, and I think "
            "we could help keep your supply chain moving with consistent, high-quality parts.\n\n"
            "Would you be open to a quick phone call this week to see if there's a fit?\n\n"
            "Feel free to call me between the hours of 7am–11am EST if that's easier.\n\n"
            "Thanks,\n"
            "August Cox\n"
            "+1 (513) 348-0483\n"
            "Sharp-Eye Precision Parts Co., Ltd."
        ),
    ),
    (
        2,
        "Follow-up #1",
        "Re: Sharp-Eye Precision Parts — following up with {{Company}}",
        (
            "Hi {{FirstName}},\n\n"
            "I sent a note over a few days back and wanted to make sure it didn't get buried.\n\n"
            "Sharp-Eye Precision Parts handles CNC turned parts at scale — swiss machining, "
            "tight tolerances, consistent output. We work with companies in {{Industry}} and "
            "I'd love to see if we can help {{Company}} the same way.\n\n"
            "No pressure — if it's not a priority right now, just let me know. "
            "But if you've got 5 minutes this week, I'm happy to talk through what we do.\n\n"
            "Please let me know if you have any questions.\n\n"
            "Thanks,\n"
            "August Cox\n"
            "+1 (513) 348-0483\n"
            "Sharp-Eye Precision Parts Co., Ltd."
        ),
    ),
    (
        3,
        "Breakup / Final Attempt",
        "Sharp-Eye Precision Parts — last note from me",
        (
            "Hi {{FirstName}},\n\n"
            "I've reached out a couple times and haven't heard back — totally understand, "
            "you're busy.\n\n"
            "I'll leave you with this: if {{Company}} ever has a bottleneck on CNC turned "
            "parts — swiss machining, tight deadlines, finicky tolerances — we're here. "
            "Sharp-Eye has been solving those kinds of problems for a long time.\n\n"
            "I won't fill up your inbox after this one. But if something changes on your end, "
            "you know where to find me.\n\n"
            "Thanks,\n"
            "August Cox\n"
            "+1 (513) 348-0483\n"
            "Sharp-Eye Precision Parts Co., Ltd."
        ),
    ),
    (
        4,
        "After Meeting",
        "Great speaking with you, {{FirstName}} — next steps",
        (
            "Hi {{FirstName}},\n\n"
            "Really appreciated the conversation earlier. Sounds like {{Company}} has some "
            "interesting projects in the pipeline.\n\n"
            "As discussed, here's what I'll do on my end:\n"
            "- {{ActionItem1}}\n"
            "- {{ActionItem2}}\n\n"
            "I'll circle back by {{FollowUpDate}} unless I hear from you sooner. "
            "Feel free to call me between the hours of 7am–11am EST with any questions.\n\n"
            "Thanks,\n"
            "August Cox\n"
            "+1 (513) 348-0483\n"
            "Sharp-Eye Precision Parts Co., Ltd."
        ),
    ),
    (
        5,
        "Quote Follow-up",
        "Sharp-Eye quote for {{Company}} — checking in",
        (
            "Hi {{FirstName}},\n\n"
            "Just following up on the quote I sent over for {{ProjectName}}. "
            "Wanted to make sure it landed safely and answer any questions you might have.\n\n"
            "Happy to walk through the numbers or adjust anything that doesn't line up "
            "with what you're expecting.\n\n"
            "Please let me know if you have any questions.\n\n"
            "Thanks,\n"
            "August Cox\n"
            "+1 (513) 348-0483\n"
            "Sharp-Eye Precision Parts Co., Ltd."
        ),
    ),
]

for i, (num, name, subject, body) in enumerate(templates, 2):
    ws5.cell(row=i, column=1, value=num)
    ws5.cell(row=i, column=2, value=name)
    ws5.cell(row=i, column=3, value=subject)
    ws5.cell(row=i, column=4, value=body)
    # Wrap body text
    ws5.cell(row=i, column=4).alignment = Alignment(vertical="top", wrap_text=True)

apply_body_style(ws5, 2, 2 + len(templates) - 1, len(TEMPLATE_COLS))
ws5.freeze_panes = "A2"

ws5.column_dimensions["A"].width = 12
ws5.column_dimensions["B"].width = 22
ws5.column_dimensions["C"].width = 50
ws5.column_dimensions["D"].width = 80

# Row heights for readability
for r in range(2, 2 + len(templates)):
    ws5.row_dimensions[r].height = 200

# ===== TAB 6: Call Scripts =====
ws6 = wb.create_sheet("Call Scripts")

SCRIPT_COLS = ["Script #", "Script Name", "Script / Notes"]
for c, name in enumerate(SCRIPT_COLS, 1):
    ws6.cell(row=1, column=c, value=name)
style_header_row(ws6, len(SCRIPT_COLS))

scripts = [
    (
        1,
        "Initial Cold Call Script",
        (
            "--- OPENING ---\n"
            "\"Hi {{FirstName}}, this is August Cox with Sharp-Eye Precision Parts. "
            "Did I catch you at an okay time?\"\n\n"
            "--- HOOK ---\n"
            "\"I'll be quick — we do CNC turned parts, swiss-type automatic lathe work "
            "mostly. I've been looking at what {{Company}} does, and I think there might "
            "be a fit on the machining side.\"\n\n"
            "--- QUALIFYING QUESTION ---\n"
            "\"Quick question — are you the right person to talk to about your machined "
            "parts supply, or should I be speaking with someone else?\"\n\n"
            "--- VALUE PROP ---\n"
            "\"We help companies in {{Industry}} get consistent, tight-tolerance turned "
            "parts without the headaches. Fast turnaround, competitive pricing, and we "
            "don't make excuses when things get tight.\"\n\n"
            "--- CALL TO ACTION ---\n"
            "\"I'd love to send over a capabilities sheet and follow up next week. "
            "What's the best email to reach you at?\"\n\n"
            "--- CLOSE ---\n"
            "\"Thanks, {{FirstName}}. I'll send that over today. "
            "Feel free to call me between 7am–11am EST if anything comes to mind.\""
        ),
    ),
    (
        2,
        "Voicemail Script",
        (
            "\"Hi {{FirstName}}, this is August Cox at Sharp-Eye Precision Parts — "
            "that's 513-348-0483. I'm reaching out because we do CNC turned parts and "
            "I'd like to see if we can help {{Company}} with your machining needs. "
            "No need to call back unless it's convenient — I'll shoot you an email as "
            "well. Thanks.\""
        ),
    ),
    (
        3,
        "Gatekeeper Bypass Tips",
        (
            "- Be polite and direct. Gatekeepers control access; treat them with respect.\n"
            "- Use their name if you can find it.\n"
            "- Don't pitch the gatekeeper — just ask for the right person.\n"
            "  Example: \"Hi, I'm trying to reach whoever handles your machined parts "
            "procurement. Can you point me in the right direction?\"\n"
            "- If they ask what it's about: \"We're a CNC machine shop and I wanted to "
            "send over some information on our capabilities. Just trying to find the "
            "right contact.\"\n"
            "- Best calling windows for manufacturers: 7am–9am EST or 4pm–5pm EST "
            "(before the floor gets busy, or after the day winds down).\n"
            "- If gatekeeper is a hard no: ask for a general email to send info. "
            "Then follow up by name next time.\n"
            "- Always log the gatekeeper's name in Notes for future calls."
        ),
    ),
    (
        4,
        "Follow-up Call Script",
        (
            "--- OPENING ---\n"
            "\"Hi {{FirstName}}, August Cox from Sharp-Eye again. I sent over some info "
            "on our CNC capabilities last week — did you get a chance to look it over?\"\n\n"
            "--- IF YES ---\n"
            "\"Great. What stood out to you? Anything you'd like me to dig into more?\"\n"
            "[Listen for pain points, timelines, current suppliers]\n\n"
            "--- IF NO ---\n"
            "\"No problem at all. The short version is: we do Swiss-type turned parts, "
            "tight tolerances, consistent quality. If you're running into any issues with "
            "your current supplier — lead times, quality, pricing — that's usually where "
            "we can help.\"\n\n"
            "--- NEXT STEP ---\n"
            "\"What would be a good time for me to check back? I don't want to be a "
            "pest, but I do think this is worth a real conversation.\"\n\n"
            "--- CLOSE ---\n"
            "\"Thanks, {{FirstName}}. I'll make a note and reach out then. "
            "Please let me know if you have any questions in the meantime.\""
        ),
    ),
]

for i, (num, name, body) in enumerate(scripts, 2):
    ws6.cell(row=i, column=1, value=num)
    ws6.cell(row=i, column=2, value=name)
    ws6.cell(row=i, column=3, value=body)
    ws6.cell(row=i, column=3).alignment = Alignment(vertical="top", wrap_text=True)

apply_body_style(ws6, 2, 2 + len(scripts) - 1, len(SCRIPT_COLS))
ws6.freeze_panes = "A2"

ws6.column_dimensions["A"].width = 12
ws6.column_dimensions["B"].width = 24
ws6.column_dimensions["C"].width = 100

for r in range(2, 2 + len(scripts)):
    ws6.row_dimensions[r].height = 280

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
OUTPUT = r"C:\Users\mapta\Documents\Claude OS\SharpEye\SharpEye_Outreach_Tracker.xlsx"
wb.save(OUTPUT)
print(f"Workbook saved to: {OUTPUT}")
